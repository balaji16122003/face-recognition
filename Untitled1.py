import cv2
import numpy as np
import os
import openpyxl
import time
from datetime import datetime

# Face recognition configuration
size = 4
haar_file = 'haarcascade_frontalface_default.xml'
datasets = 'datasets'
print('Recognizing Face. Please ensure sufficient lighting...')

# Get today's date for the file name
today_date = datetime.now().strftime('%Y-%m-%d')
excel_file_path = os.path.join(os.getcwd(), f'{today_date}.xlsx')  # Save the file in the current directory with today's date

# Function to check if the file is locked or in use
def is_file_locked(file_path):
    try:
        # Try to open the file in append mode
        with open(file_path, 'a'):
            return False  # The file is not locked
    except IOError:
        return True  # The file is locked

# Create a new workbook for today's date
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = 'Attendance'

# Add hour columns (9 AM to 8 PM)
hours = [f'{i} AM' if i < 12 else f'{i - 12} PM' for i in range(9, 21)]
sheet.append(['Name'] + hours)

# Function to check if a name exists in the sheet
def name_exists(sheet, name):
    for row in sheet.iter_rows(values_only=True):
        if name == row[0]:  # Check the first column for the name
            return True
    return False

# Function to find the correct hour column based on the current time
def get_hour_column():
    current_hour = datetime.now().hour
    if current_hour < 9:
        return None  # If before 9 AM, don't record attendance
    elif current_hour < 12:
        return current_hour - 8  # Map 9 AM to column 1
    elif current_hour < 20:
        return current_hour - 8  # Map 10 AM to column 2, ..., 8 PM to column 11
    else:
        return None  # After 8 PM, don't record attendance

# Create a list of images and a list of corresponding names
(images, labels, names, id) = ([], [], {}, 0)
for (subdirs, dirs, files) in os.walk(datasets):
    for subdir in dirs:
        names[id] = subdir
        subjectpath = os.path.join(datasets, subdir)
        for filename in os.listdir(subjectpath):
            path = subjectpath + '/' + filename
            label = id
            images.append(cv2.imread(path, 0))
            labels.append(int(label))
        id += 1

(width, height) = (130, 100)

# Create a Numpy array from the two lists above
(images, labels) = [np.array(lis) for lis in [images, labels]]

# OpenCV trains a model from the images
model = cv2.face.LBPHFaceRecognizer_create()
model.train(images, labels)

# Use face_cascade for detecting faces
face_cascade = cv2.CascadeClassifier(haar_file)
webcam = cv2.VideoCapture(0)

while True:
    (_, im) = webcam.read()
    gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, 1.3, 5)

    for (x, y, w, h) in faces:
        cv2.rectangle(im, (x, y), (x + w, y + h), (255, 0, 0), 2)
        face = gray[y:y + h, x:x + w]
        face_resize = cv2.resize(face, (width, height))

        # Try to recognize the face
        prediction = model.predict(face_resize)
        if prediction[1] < 500:
            name = '%s' % (names[prediction[0]])
            cv2.putText(im, name, (x-10, y-10), cv2.FONT_HERSHEY_PLAIN, 1, (0, 255, 0))

            # Add name to the Excel sheet if it doesn't exist
            if not name_exists(sheet, name):
                current_hour_column = get_hour_column()
                if current_hour_column is not None:
                    # Find the row for the name or add it
                    row = [name] + ['' for _ in range(12)]  # Empty placeholders for each hour
                    sheet.append(row)
                    workbook.save(excel_file_path)
                    print(f"Added '{name}' to the Excel sheet.")

            # Update the attendance for the current hour
            current_hour_column = get_hour_column()
            if current_hour_column is not None:
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    if row[0].value == name:  # Find the row with the name
                        if row[current_hour_column].value == '':  # Only add if not already marked
                            row[current_hour_column].value = 'Present'
                            workbook.save(excel_file_path)
                            print(f"Marked '{name}' as Present at {datetime.now().strftime('%H:%M:%S')}.")
                        break
        else:
            cv2.putText(im, 'not recognized', (x-10, y-10), cv2.FONT_HERSHEY_PLAIN, 1, (0, 255, 0))

    cv2.imshow('OpenCV', im)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

webcam.release()
cv2.destroyAllWindows()
