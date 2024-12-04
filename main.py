import cv2
import numpy as np
import os
import openpyxl
from datetime import datetime  # Import datetime module to get current time

# Face recognition configuration
size = 4
haar_file = 'haarcascade_frontalface_default.xml'
datasets = 'datasets'
print('Recognizing Face. Please ensure sufficient lighting...')

# Load the workbook or create a new one
excel_file_path = 'recognized_faces.xlsx'
if os.path.exists(excel_file_path):
    workbook = openpyxl.load_workbook(excel_file_path)
    if 'New' in workbook.sheetnames:
        sheet = workbook['New']
    else:
        sheet = workbook.create_sheet('New')
else:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'New'

# Function to check if a name exists in the sheet
def name_exists(sheet, name):
    for row in sheet.iter_rows(values_only=True):
        if name in row:
            return True
    return False

# Create a list of images and a list of corresponding names
(images, labels, names, id) = ([], [], {}, 0)
for (subdirs, dirs, files) in os.walk(datasets):
    for subdir in dirs:
        names[id] = subdir
        subjectpath = os.path.join(datasets, subdir)
        for filename in os.listdir(subjectpath):
            path = subjectpath + '/' + filename
            label = id
            images.append(cv2.imread(path, 0))
            labels.append(int(label))
        id += 1

(width, height) = (130, 100)

# Create a Numpy array from the two lists above
(images, labels) = [np.array(lis) for lis in [images, labels]]

# OpenCV trains a model from the images
model = cv2.face.LBPHFaceRecognizer_create()
model.train(images, labels)

# Use faceCascade on camera stream
face_cascade = cv2.CascadeClassifier(haar_file)
webcam = cv2.VideoCapture(0)

while True:
    (_, im) = webcam.read()
    gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, 1.3, 5)
    for (x, y, w, h) in faces:
        cv2.rectangle(im, (x, y), (x + w, y + h), (255, 0, 0), 2)
        face = gray[y:y + h, x:x + w]
        face_resize = cv2.resize(face, (width, height))
        
        # Try to recognize the face
        prediction = model.predict(face_resize)
        if prediction[1] < 500:  # Recognized face
            name = '%s' % (names[prediction[0]])
            cv2.putText(im, name, (x-10, y-10), cv2.FONT_HERSHEY_PLAIN, 1, (0, 255, 0))
            
            # Get the current timestamp
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Add name and timestamp to Excel sheet if it doesn't exist
            if not name_exists(sheet, name):
                sheet.append([name, current_time])
                workbook.save(excel_file_path)
                print(f"Added '{name}' with timestamp '{current_time}' to the Excel sheet.")
        else:  # Invalid face (not recognized)
            cv2.putText(im, 'not recognized', (x-10, y-10), cv2.FONT_HERSHEY_PLAIN, 1, (0, 255, 0))
            
            # Get the current timestamp for invalid face
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Log "Invalid" in Excel sheet with timestamp
            if not name_exists(sheet, "Invalid"):
                sheet.append(["Invalid", current_time])
                workbook.save(excel_file_path)
                print(f"Added 'Invalid' with timestamp '{current_time}' to the Excel sheet.")

    cv2.imshow('OpenCV', im)
    
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

webcam.release()
cv2.destroyAllWindows()
